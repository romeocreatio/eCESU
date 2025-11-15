# utils/excel_writer.py
from __future__ import annotations

import json
import os
from io import BytesIO
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter


# --------------------------------------------------------------------
# Constantes nommées (adapter si besoin)
# --------------------------------------------------------------------

# ⚠️ Attention à la casse pour le cloud (Linux = sensible à "templates" vs "Templates")
# Adapte ici en fonction du vrai chemin de ta maquette dans le repo.
EXCEL_PATH_DEFAULT = os.path.join("templates", "maquette.xlsx")

SHEET_NAME = "Formations"
TABLE_NAME = "Analyse_Globale_Formations"


# --------------------------------------------------------------------
# Helpers génériques
# --------------------------------------------------------------------

def _coerce_cell_value(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, list):
        return " • ".join(str(x) for x in v if x is not None and str(x).strip() != "")
    if isinstance(v, dict):
        try:
            return json.dumps(v, ensure_ascii=False)
        except Exception:
            return str(v)
    return str(v)


def _get_table(ws, table_name: str):
    # openpyxl >= 3.1 : ws.tables est un dict {name: Table}
    tables = getattr(ws, "tables", {})
    if isinstance(tables, dict):
        return tables.get(table_name)
    # fallback (anciens openpyxl): liste
    for t in tables:
        if getattr(t, "name", None) == table_name:
            return t
    return None


def _read_table_headers(ws, table) -> Tuple[List[str], int, int, int, int]:
    """
    Retourne (headers, min_row, max_row, min_col, max_col) d'après table.ref
    """
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers: List[str] = []
    for col in range(min_col, max_col + 1):
        headers.append(str(ws.cell(row=min_row, column=col).value or "").strip())
    return headers, min_row, max_row, min_col, max_col


def _build_row_to_write(headers: List[str], json_excel: Dict[str, Any]) -> Dict[int, str]:
    """
    Construit un mapping {col_index: value} en respectant les headers EXACTS.
    Les clés JSON non présentes dans headers sont ignorées (retournées à l’appelant via un second retour).
    """
    row_values: Dict[int, str] = {}
    # mapping header -> index
    header_to_idx = {h: i for i, h in enumerate(headers)}  # 0-based
    for h, idx0 in header_to_idx.items():
        val = json_excel.get(h)
        row_values[idx0 + 1] = _coerce_cell_value(val)  # 1-based pour openpyxl
    return row_values


def _diff_keys(headers: List[str], json_excel: Dict[str, Any]) -> Tuple[List[str], List[str]]:
    """
    Retourne (json_keys_without_column, excel_columns_without_value)
    """
    json_keys = set(json_excel.keys())
    header_set = set(headers)
    without_col = sorted([k for k in json_keys if k not in header_set])
    without_val = sorted([h for h in headers if h not in json_keys])
    return without_col, without_val


def _append_row_to_workbook(
    wb,
    json_excel: Dict[str, Any],
    sheet_name: str,
    table_name: str,
) -> Dict[str, Any]:
    """
    Logique centrale qui modifie le workbook en mémoire :
    - trouve la feuille + table nommée
    - calcule la nouvelle ligne
    - étend la table
    Ne s’occupe PAS de sauvegarder (fichier ou bytes).
    """
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Feuille '{sheet_name}' introuvable dans le workbook.")
    ws = wb[sheet_name]

    table = _get_table(ws, table_name)
    if table is None:
        raise ValueError(
            f"Table nommée '{table_name}' introuvable dans la feuille '{sheet_name}'."
        )

    headers, min_row, max_row, min_col, max_col = _read_table_headers(ws, table)

    # Construire la ligne à écrire
    row_map = _build_row_to_write(headers, json_excel)
    json_keys_without_col, excel_cols_without_val = _diff_keys(headers, json_excel)

    # Ligne d’insertion = première ligne après la table
    insert_row = max_row + 1

    # Écriture cellule par cellule (respecte l’ordre des colonnes du tableau)
    for col_idx in range(min_col, max_col + 1):
        header_idx = col_idx - min_col + 1  # 1..N aligné avec row_map
        value = row_map.get(header_idx, "")
        ws.cell(row=insert_row, column=col_idx, value=value)

    # Étendre la table d’1 ligne
    start_col_letter = get_column_letter(min_col)
    end_col_letter = get_column_letter(max_col)
    new_ref = f"{start_col_letter}{min_row}:{end_col_letter}{insert_row}"
    table.ref = new_ref

    return {
        "sheet": sheet_name,
        "table": table_name,
        "inserted_row": insert_row,
        "json_keys_without_column": json_keys_without_col,  # ignorées
        "excel_columns_without_value": excel_cols_without_val,  # restées vides
    }


# --------------------------------------------------------------------
# 1) Version "historique" : écriture sur fichier disque (local)
# --------------------------------------------------------------------

def append_json_to_named_table(
    json_excel: Dict[str, Any],
    excel_path: str = EXCEL_PATH_DEFAULT,
    sheet_name: str = SHEET_NAME,
    table_name: str = TABLE_NAME,
) -> Dict[str, Any]:
    """
    Version locale (comme avant) :
    Injecte json_excel comme nouvelle ligne dans la table Excel nommée,
    en modifiant le fichier sur disque.

    - Ne crée pas de colonnes : correspondance stricte par entête.
    - Étend la table pour inclure la nouvelle ligne.
    Retourne un dict avec infos utiles (warnings, ligne d’insertion, etc.).
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Fichier Excel introuvable: {excel_path}")

    wb = load_workbook(excel_path)

    info = _append_row_to_workbook(
        wb=wb,
        json_excel=json_excel,
        sheet_name=sheet_name,
        table_name=table_name,
    )

    wb.save(excel_path)
    wb.close()

    info["excel_path"] = excel_path
    return info


# --------------------------------------------------------------------
# 2) Version cloud / Streamlit : tout en mémoire, retour en bytes
# --------------------------------------------------------------------

def append_json_to_named_table_in_memory(
    json_excel: Dict[str, Any],
    template_bytes: bytes,
    sheet_name: str = SHEET_NAME,
    table_name: str = TABLE_NAME,
) -> Tuple[bytes, Dict[str, Any]]:
    """
    Version cloud-friendly :
    - Charge un workbook à partir de `template_bytes` (Excel modèle),
    - Ajoute une ligne dans la table nommée,
    - Retourne le fichier Excel final sous forme de bytes (prêt pour un download Streamlit).

    Ne touche pas au système de fichiers, fonctionne uniquement en mémoire :
    parfait pour Streamlit Cloud.
    """
    # Charger le workbook à partir du template en mémoire
    buffer_in = BytesIO(template_bytes)
    wb = load_workbook(buffer_in)

    info = _append_row_to_workbook(
        wb=wb,
        json_excel=json_excel,
        sheet_name=sheet_name,
        table_name=table_name,
    )

    # Sauvegarde en mémoire
    buffer_out = BytesIO()
    wb.save(buffer_out)
    wb.close()
    buffer_out.seek(0)

    excel_bytes = buffer_out.getvalue()
    info["excel_path"] = None  # pas de chemin, fichier en mémoire

    return excel_bytes, info
